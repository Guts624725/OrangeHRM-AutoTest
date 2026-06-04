"""
@Author  : 谢胜强
@Time    : 2026/5/6 13:49
@Desc    : Excel读写通用封装（基于openpyxl）
            支持：xlsx格式读取/写入、追加/覆盖、表头字典返回、数据驱动
            适配：所有自动化项目的测试数据、结果导出
"""
import os
from typing import List, Dict, Optional, Union
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from Base.baseLogger import Logger
logger = Logger("baseExcel.py").getLogger()


class ExcelHandler:
    """Excel 操作工具类（企业级通用）"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        # 文件不存在时自动创建空白 xlsx，省得每次手动建
        # 测试数据文件通常由测试人员维护，如果路径配错了，这里至少不会直接抛 FileNotFound
        if not os.path.exists(self.file_path):
            wb = Workbook()
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Excel文件不存在，已自动创建：{file_path}")

    def read_excel(self, sheet_name: str, skip_header: bool = True, as_dict: bool = False) -> Union[List[tuple], List[Dict]]:
        """
        通用Excel读取方法
        :param sheet_name: 工作表名称
        :param skip_header: 是否跳过表头（默认True）
        :param as_dict: 是否返回字典格式（True=[{"用户名":"xxx"}], False=[("xxx",)]）
        :return: 数据列表
        """
        try:
            # data_only=True 很关键：如果单元格里有公式，读取的是公式计算后的值，而不是公式字符串
            # 比如 A1 里写了 =B1+C1，不加 data_only 会读到 "=B1+C1"，加了会读到 100
            wb = load_workbook(self.file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                logger.error(f"工作表【{sheet_name}】不存在！")
                return []

            ws = wb[sheet_name]
            all_data = []

            if as_dict:
                # 第一行当表头，后面每行转成字典
                # 测试用例里用字典比元组舒服多了，可以直接 row["用户名"] 而不是 row[0]
                headers = ws.iter_rows(min_row=1, max_row=1, values_only=True).__next__()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    all_data.append(row_dict)
            else:
                start_row = 2 if skip_header else 1
                for row in ws.iter_rows(min_row=start_row, values_only=True):
                    all_data.append(row)

            wb.close()
            logger.info(f"读取Excel成功：{self.file_path} -> 工作表【{sheet_name}】，共{len(all_data)}条数据")
            return all_data

        except Exception as e:
            logger.error(f"读取Excel失败：{str(e)}")
            return []

    def write_excel(self, sheet_name: str, data_list: List[list], is_append: bool = False) -> bool:
        """
        通用Excel写入方法
        :param sheet_name: 工作表名称
        :param data_list: 二维列表数据
        :param is_append: True=追加写入 | False=覆盖当前工作表
        :return: 写入结果
        """
        try:
            wb = load_workbook(self.file_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if not is_append:
                    # 覆盖写入时清空原有数据，从第一行开始写
                    # 注意：delete_rows 会真的删掉行，不是清空内容，所以 max_row 会变小
                    ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)

            # 追加时从最后一行的下一行开始，覆盖时从第一行开始
            start_row = ws.max_row + 1 if is_append else 1
            for row_idx, row_data in enumerate(data_list, start=start_row):
                for col_idx, cell_val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_val)

            wb.save(self.file_path)
            wb.close()
            logger.info(f"写入Excel成功：{'追加' if is_append else '覆盖'} -> 工作表【{sheet_name}】")
            return True

        except Exception as e:
            logger.error(f"写入Excel失败：{str(e)}")
            return False


if __name__ == '__main__':
    excel_path = r"E:\develop\PythonProject\PythonProject\TestFramework_po\Data\DataDriver\ExcelDriver\project01_auto_test\Excel数据驱动-登录.xlsx"
    excel = ExcelHandler(excel_path)

    data = excel.read_excel("Sheet1")
    print("元组格式：", data)

    data_dict = excel.read_excel("Sheet1", as_dict=True)
    print("字典格式：", data_dict)

    excel.write_excel("Sheet1", [["5", "test", "123456", "登录成功"]], is_append=True)